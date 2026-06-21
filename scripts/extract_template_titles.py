"""Extract authoritative template titles from the EBA Annotated Table Layout.

The DPM Access DB only yields ~22/148 template titles (access-parser can't read the
memo-overflow Name fields). The EBA "Annotated Table Layout" XLSX files carry the full,
official titles in their TOC sheet (e.g. K_61.00 -> "EU KM1 - Key metrics template").

Source (download, gitignored):
  https://www.eba.europa.eu/sites/default/files/2025-07/44989a9c-e7e8-4126-8032-1156dc1c4b51/3.d%20DPM%20table%20layout%20and%20data%20point%20categorization__new.zip

Reads every "*PILLAR3*.xlsx" layout straight from that zip (CODIS, P3DH, ESG/FIN/… DIS),
merging TOC titles. Writes codebook/template_titles.csv (template, title) — small,
committed, consumed by build_codebook.py.
"""

from pathlib import Path
from io import BytesIO
import csv
import zipfile
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ZIP = ROOT / "codebook" / "dpm_table_layout.zip"
OUT = ROOT / "codebook" / "template_titles.csv"


def titles_from_workbook(data: bytes):
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    if "TOC" not in wb.sheetnames:
        return
    for row in wb["TOC"].iter_rows(values_only=True):
        vals = [c for c in row if c is not None]  # TOC has leading empty cells
        if len(vals) < 2:
            continue
        code, title = vals[0], vals[1]
        if isinstance(code, str) and code.startswith("K_") and isinstance(title, str) and title.strip():
            yield code.strip(), title.strip()


def main():
    titles = {}
    with zipfile.ZipFile(ZIP) as z:
        members = [n for n in z.namelist() if "PILLAR3" in n and n.endswith(".xlsx")]
        print(f"Reading {len(members)} PILLAR3 layout(s) from {ZIP.name}")
        for name in members:
            for code, title in titles_from_workbook(z.read(name)):
                titles.setdefault(code, title)  # first (CODIS-ordered) wins

    rows = [{"template": c, "title": t} for c, t in sorted(titles.items())]
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["template", "title"])
        w.writeheader()
        w.writerows(rows)
    print(f"✓ {OUT}  ({len(rows)} template titles)")


if __name__ == "__main__":
    main()
